#!/usr/bin/env python3
"""
NetComponents RFQ Automation Tool

Automates searching NetComponents, filtering suppliers by region,
and submitting RFQs for electronic components.

Usage:
    # Single part search
    python main.py -p "LM358N" --quantity 1000

    # Multiple parts from file
    python main.py -f parts.xlsx --quantity 500 --workers 3

    # Dry run (no actual submissions)
    python main.py -p "NE555P" --quantity 100 --dry-run

    # Review mode (single worker, step by step)
    python main.py -p "MAX232" --quantity 250 --review
"""

import argparse
import asyncio
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook

import config
from browser import BrowserSession
from search import search_part, select_suppliers, SearchResult
from rfq import submit_rfqs_for_part, RFQResult, summarize_rfq_results


# =============================================================================
# File I/O
# =============================================================================

def load_parts_from_file(filepath: str) -> list[dict]:
    """
    Load part numbers from Excel file.
    Expects columns: part_number, quantity (optional), target_price (optional)
    """
    parts = []
    path = Path(filepath)

    if not path.exists():
        print(f"Error: File not found: {filepath}")
        return parts

    if path.suffix.lower() in (".xlsx", ".xls"):
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active

        # Get header row
        headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]

        # Find column indices
        pn_idx = None
        qty_idx = None
        price_idx = None

        for i, header in enumerate(headers):
            if "part" in header or "mpn" in header:
                pn_idx = i
            elif "qty" in header or "quantity" in header:
                qty_idx = i
            elif "price" in header or "target" in header:
                price_idx = i

        if pn_idx is None:
            print("Error: Could not find part number column")
            return parts

        # Read data rows
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[pn_idx]:
                part = {
                    "part_number": str(row[pn_idx]).strip(),
                    "quantity": int(row[qty_idx]) if qty_idx and row[qty_idx] else None,
                    "target_price": float(row[price_idx]) if price_idx and row[price_idx] else None,
                }
                parts.append(part)

        wb.close()

    else:
        # Plain text file, one part per line
        with open(filepath, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    parts.append({"part_number": line, "quantity": None, "target_price": None})

    print(f"Loaded {len(parts)} parts from {filepath}")
    return parts


def save_results_to_excel(results: list[dict], output_dir: Path, filename: str = None):
    """Save RFQ results to Excel file."""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rfq_results_{timestamp}.xlsx"

    output_path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "RFQ Results"

    # Headers
    headers = ["Part Number", "Supplier", "Country", "Region", "Quantity",
               "Price", "Success", "Message", "Timestamp"]
    ws.append(headers)

    # Data rows
    for result in results:
        ws.append([
            result.get("part_number", ""),
            result.get("supplier", ""),
            result.get("country", ""),
            result.get("region", ""),
            result.get("quantity", ""),
            result.get("price", ""),
            "Yes" if result.get("success") else "No",
            result.get("message", ""),
            result.get("timestamp", ""),
        ])

    wb.save(output_path)
    print(f"Results saved to: {output_path}")
    return output_path


# =============================================================================
# Worker Functions
# =============================================================================

async def process_part(
    session: BrowserSession,
    part_number: str,
    quantity: int,
    target_price: Optional[float],
    max_suppliers: int,
    dry_run: bool,
    page=None
) -> list[dict]:
    """
    Process a single part: search, filter suppliers, submit RFQs.

    Returns list of result dicts for output.
    """
    # Use provided page or main page
    use_page = page or session.page

    results = []

    # Search for the part
    search_results = await search_part(use_page, part_number)

    if not search_results:
        print(f"  No suppliers found for {part_number}")
        return results

    # Select suppliers by region
    selected = select_suppliers(search_results, max_suppliers)
    print(f"  Selected {len(selected)} suppliers for RFQ")

    # Submit RFQs
    rfq_results = await submit_rfqs_for_part(
        page=use_page,
        part_number=part_number,
        suppliers=selected,
        quantity=quantity,
        target_price=target_price,
        dry_run=dry_run
    )

    # Convert to output format
    for i, rfq_result in enumerate(rfq_results):
        supplier_info = selected[i] if i < len(selected) else None
        results.append({
            "part_number": part_number,
            "supplier": rfq_result.supplier,
            "country": supplier_info.country if supplier_info else "",
            "region": supplier_info.region if supplier_info else "",
            "quantity": quantity,
            "price": supplier_info.price if supplier_info else "",
            "success": rfq_result.success,
            "message": rfq_result.message,
            "timestamp": rfq_result.timestamp.isoformat(),
        })

    return results


async def worker(
    worker_id: str,
    queue: asyncio.Queue,
    session: BrowserSession,
    quantity: int,
    target_price: Optional[float],
    max_suppliers: int,
    dry_run: bool,
    all_results: list
):
    """Worker coroutine that processes parts from a queue."""
    print(f"Worker {worker_id} starting...")

    # Get dedicated page for this worker
    page = await session.new_worker_page(worker_id)

    while True:
        try:
            part = queue.get_nowait()
        except asyncio.QueueEmpty:
            break

        part_number = part["part_number"]
        part_qty = part.get("quantity") or quantity
        part_price = part.get("target_price") or target_price

        print(f"[{worker_id}] Processing: {part_number}")

        try:
            results = await process_part(
                session=session,
                part_number=part_number,
                quantity=part_qty,
                target_price=part_price,
                max_suppliers=max_suppliers,
                dry_run=dry_run,
                page=page
            )
            all_results.extend(results)
        except Exception as e:
            print(f"[{worker_id}] Error processing {part_number}: {e}")
            await session.screenshot(f"error_{part_number}", page)

        queue.task_done()

    await session.close_worker_page(worker_id)
    print(f"Worker {worker_id} finished")


# =============================================================================
# Main Entry Point
# =============================================================================

async def main_async(args):
    """Async main function."""
    # Validate credentials
    if not config.validate_credentials():
        print("\nPlease set credentials in environment or .env file")
        print("See .env.example for format")
        return 1

    # Collect parts to process
    parts = []

    if args.part:
        parts.append({
            "part_number": args.part,
            "quantity": args.quantity,
            "target_price": args.price
        })

    if args.file:
        file_parts = load_parts_from_file(args.file)
        parts.extend(file_parts)

    if not parts:
        print("Error: No parts specified. Use -p or -f")
        return 1

    print(f"\nProcessing {len(parts)} part(s)")
    print(f"  Quantity: {args.quantity}")
    print(f"  Target price: {args.price or 'Not set'}")
    print(f"  Max suppliers/region: {args.max_suppliers}")
    print(f"  Workers: {args.workers}")
    print(f"  Dry run: {args.dry_run}")
    print(f"  Headless: {args.headless}")
    print()

    # Output directory
    output_dir = Path(args.output_dir) if args.output_dir else config.OUTPUT_DIR

    all_results = []

    async with BrowserSession(headless=args.headless) as session:
        # Login
        if not await session.login():
            print("Failed to login. Exiting.")
            return 1

        # Single worker mode (sequential) or review mode
        if args.workers == 1 or args.review:
            print("Running in single-worker mode...")
            for part in parts:
                part_number = part["part_number"]
                part_qty = part.get("quantity") or args.quantity
                part_price = part.get("target_price") or args.price

                results = await process_part(
                    session=session,
                    part_number=part_number,
                    quantity=part_qty,
                    target_price=part_price,
                    max_suppliers=args.max_suppliers,
                    dry_run=args.dry_run
                )
                all_results.extend(results)

                if args.review:
                    input("Press Enter to continue to next part...")

        # Parallel worker mode
        else:
            print(f"Running in parallel mode with {args.workers} workers...")

            # Create queue and populate with parts
            queue = asyncio.Queue()
            for part in parts:
                await queue.put(part)

            # Start workers with staggered timing
            worker_tasks = []
            for i in range(args.workers):
                worker_id = f"worker-{i+1}"
                task = asyncio.create_task(
                    worker(
                        worker_id=worker_id,
                        queue=queue,
                        session=session,
                        quantity=args.quantity,
                        target_price=args.price,
                        max_suppliers=args.max_suppliers,
                        dry_run=args.dry_run,
                        all_results=all_results
                    )
                )
                worker_tasks.append(task)

                # Stagger worker starts by 1 second
                if i < args.workers - 1:
                    await asyncio.sleep(1.0)

            # Wait for all workers to complete
            await asyncio.gather(*worker_tasks)

    # Save results
    if all_results:
        save_results_to_excel(all_results, output_dir)

    # Print summary
    print("\n" + "=" * 50)
    print("SUMMARY")
    print("=" * 50)
    total = len(all_results)
    successful = sum(1 for r in all_results if r.get("success"))
    print(f"Total RFQs: {total}")
    print(f"Successful: {successful}")
    print(f"Failed: {total - successful}")

    return 0


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="NetComponents RFQ Automation Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s -p "LM358N" --quantity 1000
  %(prog)s -f parts.xlsx -q 500 -w 3
  %(prog)s -p "NE555P" -q 100 --dry-run
  %(prog)s -p "MAX232" -q 250 --review
        """
    )

    # Input options
    input_group = parser.add_argument_group("Input")
    input_group.add_argument(
        "-p", "--part",
        help="Single part number to search"
    )
    input_group.add_argument(
        "-f", "--file",
        help="Excel/text file with part numbers"
    )

    # RFQ options
    rfq_group = parser.add_argument_group("RFQ Options")
    rfq_group.add_argument(
        "-q", "--quantity",
        type=int,
        default=100,
        help="Quantity to request (default: 100)"
    )
    rfq_group.add_argument(
        "--price",
        type=float,
        help="Target price per unit"
    )
    rfq_group.add_argument(
        "--max-suppliers",
        type=int,
        default=config.MAX_SUPPLIERS_PER_REGION,
        help=f"Max suppliers per region (default: {config.MAX_SUPPLIERS_PER_REGION})"
    )

    # Execution options
    exec_group = parser.add_argument_group("Execution")
    exec_group.add_argument(
        "-w", "--workers",
        type=int,
        default=1,
        choices=range(1, 11),
        metavar="N",
        help="Number of parallel workers (1-10, default: 1)"
    )
    exec_group.add_argument(
        "--dry-run",
        action="store_true",
        help="Search only, don't submit RFQs"
    )
    exec_group.add_argument(
        "--headless",
        action="store_true",
        default=True,
        help="Run browser in headless mode (default: True)"
    )
    exec_group.add_argument(
        "--no-headless",
        action="store_false",
        dest="headless",
        help="Run browser in visible mode"
    )
    exec_group.add_argument(
        "--review",
        action="store_true",
        help="Review mode: single worker, pause between parts"
    )

    # Output options
    output_group = parser.add_argument_group("Output")
    output_group.add_argument(
        "--output-dir",
        help=f"Output directory (default: {config.OUTPUT_DIR})"
    )

    args = parser.parse_args()

    # Validate input
    if not args.part and not args.file:
        parser.error("Must specify -p/--part or -f/--file")

    # Review mode forces single worker
    if args.review and args.workers > 1:
        print("Note: Review mode forces single worker")
        args.workers = 1

    # Run async main
    try:
        return asyncio.run(main_async(args))
    except KeyboardInterrupt:
        print("\nInterrupted by user")
        return 130


if __name__ == "__main__":
    sys.exit(main())
