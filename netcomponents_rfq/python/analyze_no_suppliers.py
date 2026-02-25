#!/usr/bin/env python3
"""
Analyze RFQ results to identify lines with no suppliers.

Shows whether other MPNs under the same CPC got quotes,
helping identify which CPC lines truly have "no quote" coverage.

Usage:
    python analyze_no_suppliers.py <results_excel_file> [rfq_number]

Example:
    python analyze_no_suppliers.py RFQ_1130292_Results_2026-02-25_212623.xlsx 1130292

If rfq_number is provided and CPC column is missing, it will be looked up from the database.

Output: <input_file>_NoSuppliers_Analysis.xlsx
"""

import sys
import subprocess
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def get_cpc_mapping_from_db(rfq_number):
    """Query database to get MPN -> CPC mapping for an RFQ."""
    query = f"""
    SELECT
        COALESCE(m.chuboe_mpn_clean, m.chuboe_mpn) as mpn,
        COALESCE(l.chuboe_cpc_clean, l.chuboe_cpc) as cpc
    FROM adempiere.chuboe_rfq r
    JOIN adempiere.chuboe_rfq_line l ON r.chuboe_rfq_id = l.chuboe_rfq_id
    JOIN adempiere.chuboe_rfq_line_mpn m ON l.chuboe_rfq_line_id = m.chuboe_rfq_line_id
    WHERE r.value = '{rfq_number}'
      AND l.isactive = 'Y'
      AND m.isactive = 'Y';
    """

    result = subprocess.run(
        ['psql', '-t', '-A', '-F', '|', '-c', query],
        capture_output=True,
        text=True
    )

    if result.returncode != 0:
        print(f"Database error: {result.stderr}")
        return {}

    mapping = {}
    for line in result.stdout.strip().split('\n'):
        if line and '|' in line:
            parts = line.split('|')
            mpn = parts[0].strip()
            cpc = parts[1].strip() if len(parts) > 1 else ''
            if mpn:
                mapping[mpn] = cpc

    return mapping


def analyze_no_suppliers(input_file, rfq_number=None):
    """Analyze RFQ results for no-supplier lines and CPC coverage."""

    # Read the results file
    df = pd.read_excel(input_file)

    # Check required columns
    required = ['Part Number', 'Status']
    missing = [col for col in required if col not in df.columns]
    if missing:
        print(f"Error: Missing required columns: {missing}")
        print(f"Available columns: {list(df.columns)}")
        return

    # If CPC column is missing, try to look it up from database
    if 'CPC' not in df.columns:
        if rfq_number:
            print(f"CPC column not found. Looking up from database for RFQ {rfq_number}...")
            cpc_mapping = get_cpc_mapping_from_db(rfq_number)
            if cpc_mapping:
                df['CPC'] = df['Part Number'].map(cpc_mapping).fillna('')
                print(f"  Found CPC mappings for {len(cpc_mapping)} MPNs")
            else:
                print("  Warning: Could not retrieve CPC mappings from database")
                df['CPC'] = ''
        else:
            print("Error: CPC column not found in results file.")
            print("Provide RFQ number as second argument to look up CPC from database.")
            print("Example: python analyze_no_suppliers.py results.xlsx 1130292")
            return

    # Build CPC -> MPN quote status mapping
    cpc_status = {}
    for _, row in df.iterrows():
        cpc = row.get('CPC', '')
        mpn = row.get('Part Number', '')
        status = row.get('Status', '')

        if pd.isna(cpc) or cpc == '':
            continue

        if cpc not in cpc_status:
            cpc_status[cpc] = {}

        # Track if MPN got quotes (SENT status means RFQ was sent)
        if mpn not in cpc_status[cpc]:
            cpc_status[cpc][mpn] = {'sent': False, 'line': row.get('RFQ Line', '')}

        if status == 'SENT':
            cpc_status[cpc][mpn]['sent'] = True

    # Find NO_SUPPLIERS entries and analyze CPC coverage
    no_supplier_rows = []
    seen = set()  # Avoid duplicate MPN entries

    for _, row in df.iterrows():
        status = row.get('Status', '')
        if status != 'NO_SUPPLIERS':
            continue

        cpc = row.get('CPC', '')
        mpn = row.get('Part Number', '')

        # Skip duplicates
        key = (cpc, mpn)
        if key in seen:
            continue
        seen.add(key)

        # Check if other MPNs under same CPC got quotes
        other_mpns_quoted = []
        if cpc in cpc_status:
            for other_mpn, info in cpc_status[cpc].items():
                if other_mpn != mpn and info['sent']:
                    other_mpns_quoted.append(other_mpn)

        cpc_has_quotes = len(other_mpns_quoted) > 0

        no_supplier_rows.append({
            'line_number': row.get('RFQ Line', ''),
            'cpc': cpc,
            'mpn': mpn,
            'qty': row.get('Qty Requested', ''),
            'cpc_has_quotes': 'Yes' if cpc_has_quotes else 'NO - NEEDS ATTENTION',
            'other_mpns_quoted': ', '.join(other_mpns_quoted[:5]) if other_mpns_quoted else ''
        })

    if not no_supplier_rows:
        print("No 'NO_SUPPLIERS' entries found in the results.")
        return

    # Sort by CPC coverage status (NO first, then Yes)
    no_supplier_rows.sort(key=lambda x: (x['cpc_has_quotes'] == 'Yes', x['line_number']))

    # Create output Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'No Suppliers Analysis'

    headers = ['RFQ Line', 'CPC', 'MPN (No Suppliers)', 'Qty', 'CPC Has Other Quotes?', 'Other MPNs Quoted']

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
    yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    no_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row_num = 2
    needs_attention_count = 0

    for r in no_supplier_rows:
        ws.cell(row=row_num, column=1, value=r['line_number'])
        ws.cell(row=row_num, column=2, value=r['cpc'])
        ws.cell(row=row_num, column=3, value=r['mpn'])
        ws.cell(row=row_num, column=4, value=r['qty'])
        ws.cell(row=row_num, column=5, value=r['cpc_has_quotes'])
        ws.cell(row=row_num, column=6, value=r['other_mpns_quoted'])

        # Color the "CPC Has Other Quotes?" column
        quote_cell = ws.cell(row=row_num, column=5)
        if 'Yes' in r['cpc_has_quotes']:
            quote_cell.fill = yes_fill
        else:
            quote_cell.fill = no_fill
            needs_attention_count += 1

        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    for col in range(1, 7):
        max_length = max(len(str(cell.value or '')) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

    # Generate output filename
    input_path = Path(input_file)
    output_file = input_path.parent / f"{input_path.stem}_NoSuppliers_Analysis.xlsx"

    wb.save(output_file)
    wb.close()

    # Summary
    print(f"\n{'='*60}")
    print("NO SUPPLIERS ANALYSIS")
    print(f"{'='*60}")
    print(f"Input file: {input_file}")
    print(f"Total MPNs with no suppliers: {len(no_supplier_rows)}")
    print(f"CPCs needing attention (no other quotes): {needs_attention_count}")
    print(f"CPCs covered by other MPNs: {len(no_supplier_rows) - needs_attention_count}")
    print(f"\nOutput saved to: {output_file}")
    print(f"{'='*60}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python analyze_no_suppliers.py <results_excel_file> [rfq_number]")
        print("Example: python analyze_no_suppliers.py RFQ_1130292_Results_2026-02-25_212623.xlsx 1130292")
        print("\nIf CPC column is missing from results, provide rfq_number to look it up from database.")
        sys.exit(1)

    input_file = sys.argv[1]
    rfq_number = sys.argv[2] if len(sys.argv) > 2 else None
    analyze_no_suppliers(input_file, rfq_number)
