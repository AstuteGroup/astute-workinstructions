#!/usr/bin/env python3
"""
Batch RFQ submission from system RFQ number

Pulls part numbers and quantities from an RFQ in iDempiere,
then submits NetComponents RFQs to qualifying suppliers.

Features:
- 3 parallel browser workers for faster processing
- Randomized timing jitter to appear natural
- Date code prioritization
- Quantity adjustment to encourage supplier quoting

Usage:
    python batch_rfqs_from_system.py <rfq_number>

Example:
    python batch_rfqs_from_system.py 1008627

Output file: RFQ_<rfq_number>_Results_YYYY-MM-DD_HHMMSS.xlsx
"""

import sys
import subprocess
import asyncio
import time
import re
import random
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config


def jitter_sleep(base_seconds):
    """Return sleep duration with random jitter (±40%)"""
    jitter = random.uniform(1 - config.JITTER_RANGE, 1 + config.JITTER_RANGE)
    return base_seconds * jitter


async def sleep_with_jitter(base_seconds):
    """Async sleep with randomized jitter"""
    await asyncio.sleep(jitter_sleep(base_seconds))


def get_rfq_lines_from_db(rfq_number):
    """Query database for RFQ line items (pulls from chuboe_rfq_line_mpn)"""
    query = f"""
    SELECT
        COALESCE(m.chuboe_mpn_clean, m.chuboe_mpn) as part_number,
        COALESCE(m.qty, l.qty) as quantity,
        COALESCE(m.chuboe_mfr_text, l.chuboe_mfr_text) as manufacturer,
        l.line as line_number,
        COALESCE(l.chuboe_cpc_clean, l.chuboe_cpc) as cpc
    FROM adempiere.chuboe_rfq r
    JOIN adempiere.chuboe_rfq_line l ON r.chuboe_rfq_id = l.chuboe_rfq_id
    JOIN adempiere.chuboe_rfq_line_mpn m ON l.chuboe_rfq_line_id = m.chuboe_rfq_line_id
    WHERE r.value = '{rfq_number}'
      AND l.isactive = 'Y'
      AND m.isactive = 'Y'
      AND COALESCE(m.qty, l.qty) > 0
    ORDER BY l.line, m.chuboe_rfq_line_mpn_id;
    """

    result = subprocess.run(
        ['psql', '-t', '-A', '-F', '|', '-c', query],
        capture_output=True,
        text=True
    )

    if result.returncode != 0:
        print(f"Database error: {result.stderr}")
        return []

    parts = []
    for line in result.stdout.strip().split('\n'):
        if line:
            fields = line.split('|')
            if len(fields) >= 2:
                parts.append({
                    'part_number': fields[0].strip(),
                    'quantity': int(float(fields[1])),
                    'manufacturer': fields[2].strip() if len(fields) > 2 else '',
                    'line_number': int(fields[3]) if len(fields) > 3 and fields[3] else 0,
                    'cpc': fields[4].strip() if len(fields) > 4 else ''
                })

    return parts


def create_output_excel(results, rfq_number, output_path):
    """Create output Excel file with RFQ results"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'RFQ {rfq_number} Results'

    headers = ['RFQ Line', 'CPC', 'Part Number', 'Qty Requested', 'Qty Sent', 'Supplier', 'Region',
               'Supplier Qty', 'Min Order $', 'Est Value $', 'Qualifying', 'Qual Amer', 'Qual Eur', 'Selected',
               'Status', 'Timestamp', 'Error', 'Worker']

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    omitted_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')  # Yellow for omitted

    row_num = 2
    for r in results:
        ws.cell(row=row_num, column=1, value=r.get('line_number', ''))
        ws.cell(row=row_num, column=2, value=r.get('cpc', ''))
        ws.cell(row=row_num, column=3, value=r.get('part_number', ''))
        ws.cell(row=row_num, column=4, value=r.get('qty_requested', ''))
        ws.cell(row=row_num, column=5, value=r.get('qty_sent', ''))
        ws.cell(row=row_num, column=6, value=r.get('supplier', ''))
        ws.cell(row=row_num, column=7, value=r.get('region', ''))
        ws.cell(row=row_num, column=8, value=r.get('supplier_qty', ''))
        ws.cell(row=row_num, column=9, value=r.get('min_order_value', ''))
        ws.cell(row=row_num, column=10, value=r.get('est_value', ''))
        ws.cell(row=row_num, column=11, value=r.get('qualifying_total', ''))
        ws.cell(row=row_num, column=12, value=r.get('qualifying_americas', ''))
        ws.cell(row=row_num, column=13, value=r.get('qualifying_europe', ''))
        ws.cell(row=row_num, column=14, value=r.get('selected_count', ''))
        ws.cell(row=row_num, column=15, value=r.get('status', ''))
        ws.cell(row=row_num, column=16, value=r.get('timestamp', ''))
        ws.cell(row=row_num, column=17, value=r.get('error', '') or r.get('reason', ''))
        ws.cell(row=row_num, column=18, value=r.get('worker_id', ''))

        status_cell = ws.cell(row=row_num, column=15)
        if r.get('status') == 'SENT':
            status_cell.fill = success_fill
        elif r.get('status') == 'FAILED':
            status_cell.fill = fail_fill
        elif r.get('status') == 'OMITTED':
            status_cell.fill = omitted_fill

        for col in range(1, 19):
            ws.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    for col in range(1, 19):
        max_length = max(len(str(cell.value or '')) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)

    wb.save(output_path)
    wb.close()


async def process_part(page, part_number, quantity, line_number, worker_id, cpc='', franchise_data=None):
    """Process a single part number and return results

    Args:
        franchise_data: Optional dict with 'franchise_qty', 'franchise_bulk_price' for min order filtering
    """
    results = []
    omitted_suppliers = []  # Track suppliers filtered out by min order value

    print(f'  [W{worker_id}] Searching for {part_number}...', flush=True)
    search_start = time.time()

    # Search box should already be ready (worker navigated here)
    await page.fill('#PartsSearched_0__PartNumber', part_number)
    # Wait for search button to be visible before clicking
    await page.wait_for_selector('#btnSearch', state='visible', timeout=15000)
    await page.click('#btnSearch')
    await sleep_with_jitter(7)  # Base 7 sec with jitter
    print(f'    [W{worker_id}] Search complete ({time.time() - search_start:.1f}s)', flush=True)

    rows = await page.query_selector_all('table#trv_0 tbody tr')

    supplier_data = {}
    in_stock_section = False
    current_region = 'Unknown'

    for row in rows:
        cells = await row.query_selector_all('td')
        row_text = (await row.inner_text() or '').lower()

        # Header rows have few cells (1-3), data rows have 16+
        is_header_row = len(cells) < 5

        if is_header_row:
            # Region headers
            if 'americas' in row_text:
                current_region = 'Americas'
            elif 'europe' in row_text:
                current_region = 'Europe'
            elif 'asia' in row_text or 'other' in row_text:
                current_region = 'Asia/Other'
            # Section headers
            if 'in stock' in row_text or 'in-stock' in row_text:
                in_stock_section = True
            elif 'brokered' in row_text:
                in_stock_section = False
            continue

        # Data rows - must have 16+ cells
        if len(cells) < 16:
            continue

        if not in_stock_section or current_region == 'Asia/Other':
            continue

        supplier_cell = cells[15]
        link = await supplier_cell.query_selector('a')
        if not link:
            continue
        supplier_name = (await link.inner_text()).strip()
        if not supplier_name:
            continue

        # Skip franchised/authorized distributors (marked with 'ncauth' class)
        auth_icon = await supplier_cell.query_selector('.ncauth')
        if auth_icon:
            continue

        # Get date code from column 4
        dc_text = ''
        dc_year = None
        dc_ambiguous = False
        try:
            dc_text = (await cells[4].inner_text()).strip()
            dc_year, dc_ambiguous = config.parse_date_code(dc_text)
        except Exception:
            pass

        qty = 0
        try:
            qty_text = (await cells[8].inner_text()).strip()
            match = re.match(r'^(\d+)', qty_text.replace(',', ''))
            if match:
                qty = int(match.group(1))
        except Exception:
            pass

        key = f"{supplier_name}|{current_region}"
        if key not in supplier_data:
            supplier_data[key] = {
                'name': supplier_name,
                'region': current_region,
                'total_qty': 0,
                'best_dc_year': None,
                'best_dc_text': '',
                'dc_ambiguous': False
            }
        supplier_data[key]['total_qty'] += qty

        # Keep the best (freshest) date code
        if dc_year is not None:
            if supplier_data[key]['best_dc_year'] is None or dc_year > supplier_data[key]['best_dc_year']:
                supplier_data[key]['best_dc_year'] = dc_year
                supplier_data[key]['best_dc_text'] = dc_text
                supplier_data[key]['dc_ambiguous'] = dc_ambiguous

    # Determine date code status for each supplier
    for s in supplier_data.values():
        s['dc_status'] = config.get_dc_status(s.get('best_dc_year'), s.get('dc_ambiguous', False))

    # Split by region and sort by priority score
    americas = [s for s in supplier_data.values() if s['region'] == 'Americas']
    europe = [s for s in supplier_data.values() if s['region'] == 'Europe']

    americas.sort(key=lambda x: config.supplier_priority_score(x, quantity), reverse=True)
    europe.sort(key=lambda x: config.supplier_priority_score(x, quantity), reverse=True)

    # Apply coverage-based filtering to remove tiny-qty suppliers when good coverage exists
    # Combine all suppliers first to assess overall coverage, then split back
    all_suppliers = americas + europe
    filtered_all = config.filter_by_coverage(all_suppliers, quantity)
    filtered_names = {s['name'] for s in filtered_all}

    americas = [s for s in americas if s['name'] in filtered_names]
    europe = [s for s in europe if s['name'] in filtered_names]

    # Track qualifying supplier counts (after coverage filter)
    qualifying_americas = len(americas)
    qualifying_europe = len(europe)
    qualifying_total = qualifying_americas + qualifying_europe

    # Use cross-region balancing: if one region is short, give extra slots to the other
    americas_slots, europe_slots = config.calculate_region_slots(qualifying_americas, qualifying_europe)

    selected_americas = americas[:americas_slots]
    selected_europe = europe[:europe_slots]

    # Add +1 extra if unknown DCs present (buffer for uncertainty)
    if config.should_add_extra_supplier(selected_americas) and len(americas) > americas_slots:
        selected_americas = americas[:americas_slots + 1]
    if config.should_add_extra_supplier(selected_europe) and len(europe) > europe_slots:
        selected_europe = europe[:europe_slots + 1]

    all_selected = selected_americas + selected_europe
    selected_count = len(all_selected)

    if not all_selected:
        print(f'    [W{worker_id}] No qualifying suppliers found')
        results.append({
            'line_number': line_number,
            'cpc': cpc,
            'part_number': part_number,
            'qty_requested': quantity,
            'qty_sent': '',
            'supplier': '',
            'region': '',
            'supplier_qty': '',
            'qualifying_total': 0,
            'qualifying_americas': 0,
            'qualifying_europe': 0,
            'selected_count': 0,
            'status': 'NO_SUPPLIERS',
            'timestamp': datetime.now().isoformat(),
            'error': 'No qualifying suppliers found',
            'worker_id': worker_id
        })
        return results

    print(f'    [W{worker_id}] Found {qualifying_total} qualifying, selected {selected_count}')

    for supplier in all_selected:
        supplier_start = time.time()

        # Adjust quantity if supplier has less than requested
        rfq_qty, qty_adjusted = config.adjust_rfq_quantity(quantity, supplier['total_qty'])
        qty_note = f" (adj)" if qty_adjusted else ""

        print(f'    [W{worker_id}] -> {supplier["name"]} qty:{rfq_qty}{qty_note}...')

        try:
            await page.goto(config.BASE_URL)
            await sleep_with_jitter(2)
            await page.fill('#PartsSearched_0__PartNumber', part_number)
            await page.click('#btnSearch')
            await sleep_with_jitter(5)

            supplier_link = await page.query_selector(f'a:has-text("{supplier["name"]}")')
            if not supplier_link:
                results.append({
                    'line_number': line_number,
                    'cpc': cpc,
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'qty_sent': rfq_qty,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'qualifying_total': qualifying_total,
                    'qualifying_americas': qualifying_americas,
                    'qualifying_europe': qualifying_europe,
                    'selected_count': selected_count,
                    'status': 'FAILED',
                    'timestamp': datetime.now().isoformat(),
                    'error': 'Supplier not found on re-search',
                    'worker_id': worker_id
                })
                continue

            await supplier_link.click()
            await sleep_with_jitter(3)  # Wait for supplier detail popup

            # Extract min order value from supplier detail popup
            min_order_value = await config.extract_min_order_value(page)
            supplier['min_order_value'] = min_order_value

            # Apply min order value filter if franchise data is provided
            if franchise_data and min_order_value:
                # Add customer_qty to franchise_data for the filter
                franchise_data_with_qty = {**franchise_data, 'customer_qty': quantity}
                should_skip, skip_reason, filter_details = config.should_skip_for_min_order_value(
                    supplier, franchise_data_with_qty
                )
                if should_skip:
                    print(f'      [W{worker_id}] OMITTED: {skip_reason}')
                    omitted_suppliers.append({
                        'line_number': line_number,
                        'cpc': cpc,
                        'part_number': part_number,
                        'qty_requested': quantity,
                        'supplier': supplier['name'],
                        'region': supplier['region'],
                        'supplier_qty': supplier['total_qty'],
                        'min_order_value': min_order_value,
                        'franchise_bulk_price': filter_details.get('franchise_bulk_price'),
                        'est_value': filter_details.get('est_value'),
                        'multiplier': filter_details.get('multiplier'),
                        'availability': filter_details.get('availability'),
                        'reason': skip_reason,
                        'status': 'OMITTED',
                        'timestamp': datetime.now().isoformat(),
                        'worker_id': worker_id
                    })
                    await page.keyboard.press('Escape')
                    await sleep_with_jitter(1)
                    continue

            rfq_link = await page.query_selector('a:has-text("E-Mail RFQ")')
            if not rfq_link:
                results.append({
                    'line_number': line_number,
                    'cpc': cpc,
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'qty_sent': rfq_qty,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'qualifying_total': qualifying_total,
                    'qualifying_americas': qualifying_americas,
                    'qualifying_europe': qualifying_europe,
                    'selected_count': selected_count,
                    'status': 'FAILED',
                    'timestamp': datetime.now().isoformat(),
                    'error': 'No RFQ option',
                    'worker_id': worker_id
                })
                await page.keyboard.press('Escape')
                await sleep_with_jitter(1)
                continue

            await rfq_link.click()
            await sleep_with_jitter(2)

            part_checkbox = await page.query_selector('#Parts_0__Selected')
            if part_checkbox:
                if not await part_checkbox.is_checked():
                    await part_checkbox.check()

            qty_input = await page.query_selector('#Parts_0__Quantity')
            if qty_input:
                await qty_input.click()
                await qty_input.fill(str(rfq_qty))

            if supplier['region'] == 'Europe':
                comments_field = await page.query_selector('#Comments')
                if comments_field:
                    await comments_field.fill('Please confirm country of origin.')

            await sleep_with_jitter(1)

            send_btn = await page.query_selector('input[type="button"].action-btn')
            if not send_btn:
                send_btn = await page.query_selector('input[value="Send RFQ"]')

            if send_btn and await send_btn.get_attribute('disabled') is None:
                await send_btn.click()
                await sleep_with_jitter(2.5)

                supplier_time = time.time() - supplier_start

                print(f'      [W{worker_id}] SENT ({supplier_time:.1f}s)')
                results.append({
                    'line_number': line_number,
                    'cpc': cpc,
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'qty_sent': rfq_qty,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'qualifying_total': qualifying_total,
                    'qualifying_americas': qualifying_americas,
                    'qualifying_europe': qualifying_europe,
                    'selected_count': selected_count,
                    'status': 'SENT',
                    'timestamp': datetime.now().isoformat(),
                    'error': '',
                    'worker_id': worker_id
                })
            else:
                results.append({
                    'line_number': line_number,
                    'cpc': cpc,
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'qty_sent': rfq_qty,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'qualifying_total': qualifying_total,
                    'qualifying_americas': qualifying_americas,
                    'qualifying_europe': qualifying_europe,
                    'selected_count': selected_count,
                    'status': 'FAILED',
                    'timestamp': datetime.now().isoformat(),
                    'error': 'Send button not found or disabled',
                    'worker_id': worker_id
                })

            await page.keyboard.press('Escape')
            await sleep_with_jitter(1)

        except Exception as e:
            results.append({
                'line_number': line_number,
                'cpc': cpc,
                'part_number': part_number,
                'qty_requested': quantity,
                'qty_sent': rfq_qty,
                'supplier': supplier['name'],
                'region': supplier['region'],
                'supplier_qty': supplier['total_qty'],
                'qualifying_total': qualifying_total,
                'qualifying_americas': qualifying_americas,
                'qualifying_europe': qualifying_europe,
                'selected_count': selected_count,
                'status': 'FAILED',
                'timestamp': datetime.now().isoformat(),
                'error': str(e),
                'worker_id': worker_id
            })

    # Add omitted suppliers to results for reporting
    results.extend(omitted_suppliers)

    return results


async def worker(worker_id, parts_queue, results_list, results_lock):
    """
    Worker coroutine - manages its own browser instance and processes parts from the queue.
    """
    print(f'[Worker {worker_id}] Starting...')

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1400, 'height': 1000})
        page = await context.new_page()

        try:
            # Login
            print(f'[Worker {worker_id}] Logging in...', flush=True)
            await page.goto(config.BASE_URL)
            await page.wait_for_selector('a:has-text("Login")', state='visible', timeout=15000)
            await sleep_with_jitter(2)
            await page.click('a:has-text("Login")')
            await page.wait_for_selector('#AccountNumber', state='visible', timeout=15000)
            await sleep_with_jitter(1)
            await page.fill('#AccountNumber', config.NETCOMPONENTS_ACCOUNT)
            await page.fill('#UserName', config.NETCOMPONENTS_USERNAME)
            await page.fill('#Password', config.NETCOMPONENTS_PASSWORD)
            await page.press('#Password', 'Enter')
            await sleep_with_jitter(5)

            # Navigate to search page and wait for both search box AND button
            await page.goto(config.BASE_URL)
            await page.wait_for_selector('#PartsSearched_0__PartNumber', state='visible', timeout=30000)
            await page.wait_for_selector('#btnSearch', state='visible', timeout=15000)
            print(f'[Worker {worker_id}] Logged in and ready', flush=True)

            # Process parts from queue
            while True:
                try:
                    part = parts_queue.get_nowait()
                except asyncio.QueueEmpty:
                    break

                print(f'\n[Worker {worker_id}] Processing Line {part["line_number"]}: {part["part_number"]} x {part["quantity"]:,}', flush=True)

                try:
                    # Navigate to search page before each part (ensures clean state)
                    await page.goto(config.BASE_URL)
                    await page.wait_for_selector('#PartsSearched_0__PartNumber', state='visible', timeout=15000)
                    await page.wait_for_selector('#btnSearch', state='visible', timeout=10000)

                    results = await process_part(
                        page,
                        part['part_number'],
                        part['quantity'],
                        part['line_number'],
                        worker_id,
                        part.get('cpc', '')
                    )

                    # Thread-safe append to results
                    async with results_lock:
                        results_list.extend(results)

                except Exception as e:
                    print(f'[Worker {worker_id}] ERROR on {part["part_number"]}: {e}', flush=True)
                    # Record the error but continue processing
                    async with results_lock:
                        results_list.append({
                            'line_number': part['line_number'],
                            'cpc': part.get('cpc', ''),
                            'part_number': part['part_number'],
                            'qty_requested': part['quantity'],
                            'qty_sent': '',
                            'supplier': '',
                            'region': '',
                            'supplier_qty': '',
                            'qualifying_total': '',
                            'qualifying_americas': '',
                            'qualifying_europe': '',
                            'selected_count': '',
                            'status': 'FAILED',
                            'timestamp': datetime.now().isoformat(),
                            'error': str(e),
                            'worker_id': worker_id
                        })

                # Brief pause between parts (with jitter)
                await sleep_with_jitter(1.5)

        except Exception as e:
            print(f'[Worker {worker_id}] ERROR: {e}')
            import traceback
            traceback.print_exc()
        finally:
            await browser.close()
            print(f'[Worker {worker_id}] Finished')


def check_lock_file(rfq_number):
    """Check if a lock file exists for this RFQ. Returns True if safe to proceed."""
    lock_file = Path(f'RFQ_{rfq_number}/.lock')

    if not lock_file.exists():
        return True

    # Lock file exists - check if it's stale
    try:
        with open(lock_file, 'r') as f:
            content = f.read().strip()
            pid = int(content.split('\n')[0]) if content else 0

        # Check if process is still running
        import os
        try:
            os.kill(pid, 0)  # Doesn't kill, just checks if process exists
            # Process is still running
            print(f'ERROR: Batch already running for RFQ {rfq_number} (PID {pid})')
            print(f'Lock file: {lock_file}')
            print('If you believe this is stale, delete the lock file and retry.')
            return False
        except OSError:
            # Process not running - stale lock file
            print(f'Warning: Stale lock file found (PID {pid} not running). Removing...')
            lock_file.unlink()
            return True
    except Exception as e:
        print(f'Warning: Could not read lock file: {e}. Removing...')
        lock_file.unlink()
        return True


def create_lock_file(rfq_number):
    """Create a lock file for this RFQ."""
    import os
    lock_file = Path(f'RFQ_{rfq_number}/.lock')
    lock_file.parent.mkdir(exist_ok=True)
    with open(lock_file, 'w') as f:
        f.write(f'{os.getpid()}\n')
        f.write(f'Started: {datetime.now().isoformat()}\n')
    return lock_file


def remove_lock_file(rfq_number):
    """Remove the lock file for this RFQ."""
    lock_file = Path(f'RFQ_{rfq_number}/.lock')
    if lock_file.exists():
        lock_file.unlink()


async def main():
    if len(sys.argv) < 2:
        print('Usage: python batch_rfqs_from_system.py <rfq_number>')
        print('Example: python batch_rfqs_from_system.py 1008627')
        sys.exit(1)

    rfq_number = sys.argv[1]

    # Check for existing lock file (prevent duplicate runs)
    if not check_lock_file(rfq_number):
        sys.exit(1)

    # Create lock file
    lock_file = create_lock_file(rfq_number)
    print(f'Lock file created: {lock_file}')

    try:
        print(f'Fetching RFQ {rfq_number} from database...')
        parts = get_rfq_lines_from_db(rfq_number)

        if not parts:
            print(f'No line items found for RFQ {rfq_number}')
            remove_lock_file(rfq_number)
            sys.exit(1)

        print(f'Found {len(parts)} line items:\n')
        for p in parts:
            print(f"  Line {p['line_number']}: {p['part_number']} x {p['quantity']:,}")

        # Create RFQ subfolder
        rfq_folder = Path(f'RFQ_{rfq_number}')
        rfq_folder.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        output_file = rfq_folder / f'Results_{timestamp}.xlsx'

        print('\n' + '=' * 60)
        print(f'NetComponents Batch RFQ Submission')
        print(f'RFQ: {rfq_number}')
        print(f'Parts to process: {len(parts)}')
        print(f'Parallel workers: {config.NUM_WORKERS}')
        print(f'Timing jitter: ±{int(config.JITTER_RANGE * 100)}%')
        print(f'Output folder: {rfq_folder}/')
        print(f'Output file: {output_file}')
        print('=' * 60)

        # Create queue and populate with parts
        parts_queue = asyncio.Queue()
        for part in parts:
            await parts_queue.put(part)

        # Shared results list with lock
        results_list = []
        results_lock = asyncio.Lock()

        start_time = time.time()

        # Launch workers
        workers = [
            worker(i + 1, parts_queue, results_list, results_lock)
            for i in range(config.NUM_WORKERS)
        ]

        # Wait for all workers to complete
        await asyncio.gather(*workers)

        # Sort results by line number for output
        results_list.sort(key=lambda x: (x.get('line_number', 0), x.get('timestamp', '')))

        print(f'\n\nWriting results to {output_file}...')
        create_output_excel(results_list, rfq_number, output_file)

        total_time = time.time() - start_time
        sent_count = len([r for r in results_list if r['status'] == 'SENT'])
        failed_count = len([r for r in results_list if r['status'] == 'FAILED'])
        no_suppliers = len([r for r in results_list if r['status'] == 'NO_SUPPLIERS'])
        omitted_count = len([r for r in results_list if r['status'] == 'OMITTED'])

        # Supplier distribution analysis
        supplier_counts = {}
        for r in results_list:
            if r['status'] == 'SENT':
                supplier = r.get('supplier', 'Unknown')
                supplier_counts[supplier] = supplier_counts.get(supplier, 0) + 1

        # Sort by count descending
        top_suppliers = sorted(supplier_counts.items(), key=lambda x: x[1], reverse=True)

        print('\n' + '=' * 60)
        print('BATCH SUMMARY')
        print('=' * 60)
        print(f'Total parts processed: {len(parts)}')
        print(f'RFQs sent: {sent_count}')
        print(f'Failed: {failed_count}')
        print(f'Omitted (min order filter): {omitted_count}')
        print(f'No suppliers found: {no_suppliers}')
        print(f'Total time: {total_time:.1f}s ({total_time/60:.1f} min)')
        print(f'Avg time per part: {total_time/len(parts):.1f}s')
        print(f'Results saved to: {output_file}')
        print('-' * 60)
        print('SUPPLIER DISTRIBUTION')
        print(f'Unique suppliers used: {len(supplier_counts)}')
        if top_suppliers:
            print('Top 10 suppliers:')
            for supplier, count in top_suppliers[:10]:
                pct = count / sent_count * 100 if sent_count > 0 else 0
                print(f'  {supplier}: {count} RFQs ({pct:.1f}%)')
        print('=' * 60)

    finally:
        # Always remove lock file when done
        remove_lock_file(rfq_number)
        print(f'Lock file removed.')


if __name__ == '__main__':
    asyncio.run(main())
