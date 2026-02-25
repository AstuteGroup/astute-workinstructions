#!/usr/bin/env python3
"""
Batch RFQ submission with Excel input/output

Input: Excel file with columns:
  - Part Number (required)
  - Quantity (required)

Output: Excel file with RFQ results:
  - Part Number
  - Quantity
  - Supplier
  - Region
  - Supplier Qty
  - Status
  - Timestamp
  - Error (if any)

Usage:
    python batch_rfqs.py <input_excel>

Example:
    python batch_rfqs.py rfq_input.xlsx

Output file is named: RFQ_Results_YYYY-MM-DD_HHMMSS.xlsx
"""

import sys
import asyncio
import time
import re
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config


def read_input_excel(filepath):
    """Read part numbers and quantities from Excel file"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    parts = []
    headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]

    # Find columns
    pn_col = None
    qty_col = None
    for i, h in enumerate(headers):
        if 'part' in h and 'number' in h:
            pn_col = i
        elif h == 'pn' or h == 'part':
            pn_col = i
        elif 'qty' in h or 'quantity' in h:
            qty_col = i

    if pn_col is None:
        # Try first column
        pn_col = 0
    if qty_col is None:
        # Try second column
        qty_col = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[pn_col] and row[qty_col]:
            try:
                qty = int(row[qty_col])
                parts.append({
                    'part_number': str(row[pn_col]).strip(),
                    'quantity': qty
                })
            except (ValueError, TypeError):
                continue

    wb.close()
    return parts


def create_output_excel(results, output_path):
    """Create output Excel file with RFQ results"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RFQ Results'

    # Headers
    headers = ['Part Number', 'Qty Requested', 'Supplier', 'Region', 'Supplier Qty',
               'Status', 'Timestamp', 'Error']

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    row_num = 2
    for r in results:
        ws.cell(row=row_num, column=1, value=r.get('part_number', ''))
        ws.cell(row=row_num, column=2, value=r.get('qty_requested', ''))
        ws.cell(row=row_num, column=3, value=r.get('supplier', ''))
        ws.cell(row=row_num, column=4, value=r.get('region', ''))
        ws.cell(row=row_num, column=5, value=r.get('supplier_qty', ''))
        ws.cell(row=row_num, column=6, value=r.get('status', ''))
        ws.cell(row=row_num, column=7, value=r.get('timestamp', ''))
        ws.cell(row=row_num, column=8, value=r.get('error', ''))

        # Color code status
        status_cell = ws.cell(row=row_num, column=6)
        if r.get('status') == 'SENT':
            status_cell.fill = success_fill
        elif r.get('status') == 'FAILED':
            status_cell.fill = fail_fill

        # Apply borders
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = thin_border

        row_num += 1

    # Auto-width columns
    for col in range(1, 9):
        max_length = max(len(str(cell.value or '')) for cell in ws[get_column_letter(col)])
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)

    wb.save(output_path)
    wb.close()


async def process_part(page, part_number, quantity, timing_data):
    """Process a single part number and return results"""
    results = []

    print(f'\n  Searching for {part_number}...')
    search_start = time.time()
    await page.fill('#PartsSearched_0__PartNumber', part_number)
    await page.click('#btnSearch')
    await asyncio.sleep(8)
    print(f'    Search complete ({time.time() - search_start:.1f}s)')

    # Parse suppliers
    rows = await page.query_selector_all('table#trv_0 tbody tr')

    supplier_data = {}
    in_stock_section = False
    current_region = 'Unknown'

    for row in rows:
        row_text = (await row.inner_text() or '').lower()

        # Track region headers
        if 'americas' in row_text and 'inventory' not in row_text and len(row_text) < 50:
            current_region = 'Americas'
            continue
        if 'europe' in row_text and 'inventory' not in row_text and len(row_text) < 50:
            current_region = 'Europe'
            continue
        if ('asia' in row_text or 'other' in row_text) and 'inventory' not in row_text and len(row_text) < 50:
            current_region = 'Asia/Other'
            continue

        # Check for section subheader rows
        if (row_text.startswith('in stock') or row_text.startswith('in-stock')) and len(row_text) < 100:
            in_stock_section = True
            continue
        if (row_text.startswith('brokered inventory') or row_text.startswith('brokered')) and len(row_text) < 100:
            in_stock_section = False
            continue

        if not in_stock_section or current_region == 'Asia/Other':
            continue

        cells = await row.query_selector_all('td')
        if len(cells) < 16:
            continue

        supplier_cell = cells[15]
        link = await supplier_cell.query_selector('a')
        if not link:
            continue
        supplier_name = (await link.inner_text()).strip()
        if not supplier_name:
            continue

        # Skip franchised/authorized distributors (marked with 'ncauth' class)
        auth_icon = await supplier_cell.query_selector('.ncauth')
        if auth_icon:
            continue

        qty = 0
        try:
            qty_text = (await cells[8].inner_text()).strip()
            match = re.match(r'^(\d+)', qty_text.replace(',', ''))
            if match:
                qty = int(match.group(1))
        except Exception:
            pass

        key = f"{supplier_name}|{current_region}"
        if key not in supplier_data:
            supplier_data[key] = {'name': supplier_name, 'region': current_region, 'total_qty': 0}
        supplier_data[key]['total_qty'] += qty

    # Select suppliers
    all_suppliers = sorted(supplier_data.values(), key=lambda x: x['total_qty'], reverse=True)
    americas = [s for s in all_suppliers if s['region'] == 'Americas']
    europe = [s for s in all_suppliers if s['region'] == 'Europe']

    americas_meet_qty = [s for s in americas if s['total_qty'] >= quantity]
    europe_meet_qty = [s for s in europe if s['total_qty'] >= quantity]

    selected_americas = americas_meet_qty[:config.MAX_SUPPLIERS_PER_REGION] if americas_meet_qty else americas[:config.MAX_SUPPLIERS_PER_REGION]
    selected_europe = europe_meet_qty[:config.MAX_SUPPLIERS_PER_REGION] if europe_meet_qty else europe[:config.MAX_SUPPLIERS_PER_REGION]

    all_selected = selected_americas + selected_europe

    if not all_selected:
        print(f'    No qualifying suppliers found')
        results.append({
            'part_number': part_number,
            'qty_requested': quantity,
            'supplier': '',
            'region': '',
            'supplier_qty': '',
            'status': 'NO_SUPPLIERS',
            'timestamp': datetime.now().isoformat(),
            'error': 'No qualifying suppliers found'
        })
        return results

    print(f'    Found {len(all_selected)} suppliers')

    # Submit RFQs
    for supplier in all_selected:
        supplier_start = time.time()
        print(f'    Submitting to {supplier["name"]}...')

        try:
            await page.goto(config.BASE_URL)
            await asyncio.sleep(2)
            await page.fill('#PartsSearched_0__PartNumber', part_number)
            await page.click('#btnSearch')
            await asyncio.sleep(6)

            supplier_link = await page.query_selector(f'a:has-text("{supplier["name"]}")')
            if not supplier_link:
                results.append({
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'status': 'FAILED',
                    'timestamp': datetime.now().isoformat(),
                    'error': 'Supplier not found on re-search'
                })
                continue

            await supplier_link.click()
            await asyncio.sleep(2)

            rfq_link = await page.query_selector('a:has-text("E-Mail RFQ")')
            if not rfq_link:
                results.append({
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'status': 'FAILED',
                    'timestamp': datetime.now().isoformat(),
                    'error': 'No RFQ option'
                })
                await page.keyboard.press('Escape')
                await asyncio.sleep(1)
                continue

            await rfq_link.click()
            await asyncio.sleep(2)

            # Fill form
            part_checkbox = await page.query_selector('#Parts_0__Selected')
            if part_checkbox:
                if not await part_checkbox.is_checked():
                    await part_checkbox.check()

            qty_input = await page.query_selector('#Parts_0__Quantity')
            if qty_input:
                await qty_input.click()
                await qty_input.fill(str(quantity))

            if supplier['region'] == 'Europe':
                comments_field = await page.query_selector('#Comments')
                if comments_field:
                    await comments_field.fill('Please confirm country of origin.')

            await asyncio.sleep(1)

            send_btn = await page.query_selector('input[type="button"].action-btn')
            if not send_btn:
                send_btn = await page.query_selector('input[value="Send RFQ"]')

            if send_btn and await send_btn.get_attribute('disabled') is None:
                await send_btn.click()
                await asyncio.sleep(3)

                supplier_time = time.time() - supplier_start
                timing_data['suppliers'].append({'name': supplier['name'], 'time': supplier_time})

                print(f'      SENT ({supplier_time:.1f}s)')
                results.append({
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'status': 'SENT',
                    'timestamp': datetime.now().isoformat(),
                    'error': ''
                })
            else:
                results.append({
                    'part_number': part_number,
                    'qty_requested': quantity,
                    'supplier': supplier['name'],
                    'region': supplier['region'],
                    'supplier_qty': supplier['total_qty'],
                    'status': 'FAILED',
                    'timestamp': datetime.now().isoformat(),
                    'error': 'Send button not found or disabled'
                })

            await page.keyboard.press('Escape')
            await asyncio.sleep(1)

        except Exception as e:
            results.append({
                'part_number': part_number,
                'qty_requested': quantity,
                'supplier': supplier['name'],
                'region': supplier['region'],
                'supplier_qty': supplier['total_qty'],
                'status': 'FAILED',
                'timestamp': datetime.now().isoformat(),
                'error': str(e)
            })

    return results


async def main():
    if len(sys.argv) < 2:
        print('Usage: python batch_rfqs.py <input_excel>')
        print('Example: python batch_rfqs.py rfq_input.xlsx')
        sys.exit(1)

    input_file = Path(sys.argv[1])
    if not input_file.exists():
        print(f'Error: Input file not found: {input_file}')
        sys.exit(1)

    # Read input
    print(f'Reading input from {input_file}...')
    parts = read_input_excel(input_file)
    print(f'Found {len(parts)} parts to process\n')

    if not parts:
        print('No parts found in input file')
        sys.exit(1)

    # Output filename with timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    output_file = Path(f'RFQ_Results_{timestamp}.xlsx')

    print('=' * 50)
    print('NetComponents Batch RFQ Submission')
    print(f'Parts to process: {len(parts)}')
    print(f'Output file: {output_file}')
    print('=' * 50)

    all_results = []
    timing_data = {'suppliers': []}
    start_time = time.time()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1400, 'height': 1000})
        page = await context.new_page()

        try:
            # Login
            print('\nLogging in...')
            login_start = time.time()
            await page.goto(config.BASE_URL)
            await asyncio.sleep(2)
            await page.click('a:has-text("Login")')
            await asyncio.sleep(2)
            await page.fill('#AccountNumber', config.NETCOMPONENTS_ACCOUNT)
            await page.fill('#UserName', config.NETCOMPONENTS_USERNAME)
            await page.fill('#Password', config.NETCOMPONENTS_PASSWORD)
            await page.press('#Password', 'Enter')
            await asyncio.sleep(5)
            print(f'  Done ({time.time() - login_start:.1f}s)')

            # Process each part
            for i, part in enumerate(parts):
                print(f'\n[{i + 1}/{len(parts)}] Processing {part["part_number"]} x {part["quantity"]:,}')
                results = await process_part(page, part['part_number'], part['quantity'], timing_data)
                all_results.extend(results)

        except Exception as e:
            print(f'\nFATAL ERROR: {e}')
            import traceback
            traceback.print_exc()
        finally:
            await browser.close()

    # Create output Excel
    print(f'\n\nWriting results to {output_file}...')
    create_output_excel(all_results, output_file)

    # Summary
    total_time = time.time() - start_time
    sent_count = len([r for r in all_results if r['status'] == 'SENT'])

    print('\n' + '=' * 50)
    print('BATCH SUMMARY')
    print('=' * 50)
    print(f'Parts processed: {len(parts)}')
    print(f'RFQs sent: {sent_count}')
    print(f'RFQs failed: {len(all_results) - sent_count}')
    print(f'Total runtime: {total_time:.1f}s ({total_time / 60:.1f} min)')
    if timing_data['suppliers']:
        avg_time = sum(s['time'] for s in timing_data['suppliers']) / len(timing_data['suppliers'])
        print(f'Avg per supplier: {avg_time:.1f}s')
    print(f'\nResults saved to: {output_file}')


if __name__ == '__main__':
    asyncio.run(main())
