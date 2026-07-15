#!/usr/bin/env python3
"""
Batch List Suppliers - NetComponents Inventory Check (READ-ONLY)

Searches NetComponents for multiple parts and outputs supplier availability.
**THIS SCRIPT DOES NOT SEND RFQs** - it only retrieves inventory data.

Usage:
    python batch_list_suppliers.py <input_file>

Input: Text file with one MPN per line, or Excel with 'part_number' and 'quantity' columns
Output: Excel file with supplier inventory details
"""

import sys
import asyncio
import re
from datetime import datetime
from pathlib import Path
from playwright.async_api import async_playwright
import config

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


async def search_part(page, part_number, min_qty=100):
    """
    Search NetComponents for a part and return supplier data.
    READ-ONLY - does not submit RFQs.
    """
    suppliers = []

    try:
        # Navigate to homepage to ensure clean search state
        await page.goto(config.BASE_URL)
        await asyncio.sleep(2)

        # Fill search and submit
        await page.fill('#PartsSearched_0__PartNumber', part_number)
        await page.click('#btnSearch')
        await asyncio.sleep(6)  # Wait for results

        # Parse results table
        rows = await page.query_selector_all('table#trv_0 tbody tr')

        in_stock_section = False
        current_region = 'Unknown'

        for row in rows:
            cells = await row.query_selector_all('td')
            row_text = (await row.inner_text() or '').lower()

            # Header rows have few cells
            is_header_row = len(cells) < 5

            if is_header_row:
                if 'americas' in row_text:
                    current_region = 'Americas'
                elif 'europe' in row_text:
                    current_region = 'Europe'
                elif 'asia' in row_text or 'other' in row_text:
                    current_region = 'Asia/Other'
                if 'in stock' in row_text or 'in-stock' in row_text:
                    in_stock_section = True
                elif 'brokered' in row_text:
                    in_stock_section = False
                continue

            # Data rows need 16+ cells
            if len(cells) < 16:
                continue

            # Skip if not in-stock or Asia/Other
            if not in_stock_section:
                continue
            if current_region == 'Asia/Other':
                continue

            # Get supplier name from column 15
            supplier_cell = cells[15]
            link = await supplier_cell.query_selector('a')
            if not link:
                continue
            supplier_name = (await link.inner_text()).strip()
            if not supplier_name:
                continue

            # Skip franchised distributors (marked with 'ncauth' class)
            auth_icon = await supplier_cell.query_selector('.ncauth')
            if auth_icon:
                continue

            # Get offered MPN from column 0
            offered_mpn = ''
            try:
                offered_mpn = (await cells[0].inner_text()).strip()
            except:
                pass

            # Get manufacturer from column 3
            mfr = ''
            try:
                mfr = (await cells[3].inner_text()).strip()
            except:
                pass

            # Get date code from column 4
            dc_text = ''
            try:
                dc_text = (await cells[4].inner_text()).strip()
            except:
                pass

            # Get description from column 5
            description = ''
            try:
                description = (await cells[5].inner_text()).strip()
            except:
                pass

            # Get country from column 7
            country = ''
            try:
                country = (await cells[7].inner_text()).strip()
            except:
                pass

            # Get quantity from column 8
            qty = 0
            try:
                qty_text = (await cells[8].inner_text()).strip()
                qty_clean = qty_text.replace(',', '')
                match = re.match(r'^(\d+)', qty_clean)
                if match:
                    qty = int(match.group(1))
            except:
                pass

            suppliers.append({
                'supplier': supplier_name,
                'region': current_region,
                'offered_mpn': offered_mpn,
                'mfr': mfr,
                'qty': qty,
                'date_code': dc_text,
                'country': country,
                'description': description[:100] if description else ''
            })

    except Exception as e:
        print(f"    Error searching {part_number}: {e}")

    return suppliers


async def main():
    if len(sys.argv) < 2:
        print('Usage: python batch_list_suppliers.py <input_file>')
        print('')
        print('Input: Text file (one MPN per line) or Excel (part_number, quantity columns)')
        print('Output: Excel with supplier inventory details')
        print('')
        print('NOTE: This script does NOT send RFQs - read-only inventory check.')
        sys.exit(1)

    input_file = Path(sys.argv[1])
    if not input_file.exists():
        print(f"ERROR: Input file not found: {input_file}")
        sys.exit(1)

    # Load parts list
    parts = []
    if input_file.suffix.lower() == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(input_file)
        ws = wb.active
        headers = [str(c.value).lower() if c.value else '' for c in ws[1]]

        mpn_col = None
        qty_col = None
        for i, h in enumerate(headers):
            if 'part' in h or 'mpn' in h:
                mpn_col = i
            if 'qty' in h or 'quantity' in h:
                qty_col = i

        if mpn_col is None:
            print("ERROR: Excel must have a column with 'part' or 'mpn' in header")
            sys.exit(1)

        for row in ws.iter_rows(min_row=2, values_only=True):
            mpn = str(row[mpn_col]).strip() if row[mpn_col] else ''
            qty = int(row[qty_col]) if qty_col and row[qty_col] else 100
            if mpn and mpn.lower() not in ('none', 'nan', ''):
                parts.append({'mpn': mpn, 'qty': qty})
    else:
        # Text file - one MPN per line
        with open(input_file) as f:
            for line in f:
                mpn = line.strip()
                if mpn and not mpn.startswith('#'):
                    parts.append({'mpn': mpn, 'qty': 100})

    # Deduplicate
    seen = set()
    unique_parts = []
    for p in parts:
        if p['mpn'] not in seen:
            seen.add(p['mpn'])
            unique_parts.append(p)
    parts = unique_parts

    print('=' * 60, flush=True)
    print('NetComponents Batch Inventory Check (READ-ONLY)', flush=True)
    print('=' * 60, flush=True)
    print(f'Input file: {input_file}', flush=True)
    print(f'Parts to search: {len(parts)}', flush=True)
    print(f'Started: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', flush=True)
    print(flush=True)
    print('NOTE: This script does NOT send RFQs - inventory check only.', flush=True)
    print('=' * 60, flush=True)
    print(flush=True)

    # Results storage
    all_results = []
    parts_with_suppliers = 0
    parts_no_suppliers = 0

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(viewport={'width': 1400, 'height': 1000})
        page = await context.new_page()

        try:
            # Login
            print('Logging in to NetComponents...', flush=True)
            await page.goto(config.BASE_URL)
            await asyncio.sleep(2)
            await page.click('a:has-text("Login")')
            await asyncio.sleep(2)
            await page.fill('#AccountNumber', config.NETCOMPONENTS_ACCOUNT)
            await page.fill('#UserName', config.NETCOMPONENTS_USERNAME)
            await page.fill('#Password', config.NETCOMPONENTS_PASSWORD)
            await page.press('#Password', 'Enter')
            await asyncio.sleep(5)
            print('  Logged in successfully.\n', flush=True)

            # Search each part
            for i, part in enumerate(parts):
                mpn = part['mpn']
                qty = part['qty']

                print(f'[{i+1}/{len(parts)}] Searching: {mpn}', flush=True)

                suppliers = await search_part(page, mpn, qty)

                if suppliers:
                    parts_with_suppliers += 1
                    # Aggregate by supplier for summary
                    by_supplier = {}
                    for s in suppliers:
                        key = s['supplier']
                        if key not in by_supplier:
                            by_supplier[key] = {'total_qty': 0, 'regions': set()}
                        by_supplier[key]['total_qty'] += s['qty']
                        by_supplier[key]['regions'].add(s['region'])

                    print(f'    Found {len(suppliers)} listings from {len(by_supplier)} suppliers', flush=True)

                    for s in suppliers:
                        all_results.append({
                            'mpn_searched': mpn,
                            'qty_needed': qty,
                            **s
                        })
                else:
                    parts_no_suppliers += 1
                    print(f'    No in-stock suppliers found', flush=True)
                    all_results.append({
                        'mpn_searched': mpn,
                        'qty_needed': qty,
                        'supplier': '(NO SUPPLIERS)',
                        'region': '',
                        'offered_mpn': '',
                        'mfr': '',
                        'qty': 0,
                        'date_code': '',
                        'country': '',
                        'description': ''
                    })

                # Brief pause between searches
                await asyncio.sleep(1)

        except Exception as e:
            print(f'ERROR: {e}')
            import traceback
            traceback.print_exc()
        finally:
            await browser.close()

    # Output to Excel
    print()
    print('=' * 60)
    print('Creating Excel output...')

    wb = Workbook()
    ws = wb.active
    ws.title = 'NC Inventory Check'

    # Headers
    headers = ['MPN Searched', 'Qty Needed', 'Supplier', 'Region', 'Offered MPN',
               'MFR', 'Supplier Qty', 'Date Code', 'Country', 'Description']
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    no_supplier_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for r in all_results:
        row = [
            r['mpn_searched'],
            r['qty_needed'],
            r['supplier'],
            r['region'],
            r['offered_mpn'],
            r['mfr'],
            r['qty'],
            r['date_code'],
            r['country'],
            r['description']
        ]
        ws.append(row)

        # Highlight no-supplier rows
        if r['supplier'] == '(NO SUPPLIERS)':
            for cell in ws[ws.max_row]:
                cell.fill = no_supplier_fill

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 40

    # Freeze header
    ws.freeze_panes = 'A2'

    # Summary sheet
    ws_summary = wb.create_sheet('Summary')
    ws_summary.append(['NetComponents Inventory Check Summary'])
    ws_summary.append([])
    ws_summary.append(['Total MPNs searched:', len(parts)])
    ws_summary.append(['MPNs with suppliers:', parts_with_suppliers])
    ws_summary.append(['MPNs without suppliers:', parts_no_suppliers])
    ws_summary.append(['Total supplier listings:', len([r for r in all_results if r['supplier'] != '(NO SUPPLIERS)'])])
    ws_summary.append([])
    ws_summary.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S CT')])
    ws_summary.append([])
    ws_summary.append(['NOTE: This is an inventory CHECK - no RFQs were sent.'])

    # Save
    output_file = input_file.parent / f'NC_Inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(output_file)

    print(f'Output saved to: {output_file}')
    print()
    print('=' * 60)
    print('SUMMARY')
    print('=' * 60)
    print(f'Total MPNs searched: {len(parts)}')
    print(f'MPNs with suppliers: {parts_with_suppliers}')
    print(f'MPNs without suppliers: {parts_no_suppliers}')
    print(f'Total supplier listings: {len([r for r in all_results if r["supplier"] != "(NO SUPPLIERS)"])}')
    print()
    print('NOTE: This was an inventory CHECK only - NO RFQs were sent.')
    print('=' * 60)


if __name__ == '__main__':
    asyncio.run(main())
